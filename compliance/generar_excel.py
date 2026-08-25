# -*- coding: utf-8 -*-
r"""
generar_excel.py

La matriz de riesgos penales, en Excel y con la metodología nueva dentro.

Ignacio, 25/08/2026: "no veo la matriz de riesgos mejorada en Excel". Aquí
está: un .xlsx que se abre, se rellena y calcula solo, sin depender de IA
ELECE. La herramienta lo genera y lo lee, pero el fichero se defiende solo
delante de un cliente o de un auditor.

QUÉ CAMBIA RESPECTO A LA MATRIZ VIEJA, en la propia hoja de cálculo:

  EL RIESGO ES PRODUCTO.  P x I, de 1 a 25. Se acabó que probabilidad 5 con
                          impacto 1 valga lo mismo que 1 con 5.
  LA PROBABILIDAD SE ARGUMENTA.  Tres factores objetivos -exposición,
                          frecuencia del delito, historial de la empresa- y
                          dos reglas que se explican en una frase.
  LOS CONTROLES MULTIPLICAN.  Residual = inherente x PRODUCTO(1-eficacia).
                          En Excel eso se hace con EXP(SUMAR.SI(...LN...)),
                          así que sigue siendo SUMAR.SI: la misma función
                          que ya usaban, pero multiplicando.
  LOS RANGOS NO SE QUEDAN CORTOS.  El fallo que dejó 93 controles sin contar
                          en la matriz anterior venía de un SUMIF con el
                          rango escrito a mano. Aquí los rangos son de
                          columna entera: añadir filas no rompe nada.
  EL DELITO LO MANDA SU PEOR CONDUCTA, no la media.
  EL RESUMEN DA DISTRIBUCIÓN, no medias. Nadie gestiona un riesgo medio.

Se genera con openpyxl, que ya está instalado.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import catalogo as cat                               # noqa: E402
import metodologia as me                             # noqa: E402

ROJO = "8A2742"
CREMA = "F3EFEA"
LINEA = Side(style="thin", color="E0DAD1")
BORDE = Border(bottom=LINEA)

# Cuántas filas se dejan preparadas con fórmulas para escribir encima.
FILAS_MATRIZ = 300
FILAS_CONTROLES = 800


def _cabecera(ws, fila, titulos, anchos=None):
    for i, t in enumerate(titulos, start=1):
        c = ws.cell(fila, i, t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=ROJO)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 30
    for i, a in enumerate(anchos or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = ws.cell(fila + 1, 1)


def _titulo(ws, texto, sub=""):
    ws["A1"] = texto
    ws["A1"].font = Font(bold=True, size=15, color=ROJO)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(size=10, color="464237", italic=True)


# ---------------------------------------------------------------------------
def _hoja_criterios(wb):
    """Las escalas ancladas. Es la hoja que se enseña cuando preguntan
    'y esto por qué es un 3'."""
    ws = wb.create_sheet("CRITERIOS")
    _titulo(ws, "Criterios de valoración",
            "Cada nivel con su definición y la evidencia que lo demuestra. "
            "Un número sin criterio detrás es una opinión.")
    fila = 4
    bloques = [
        ("PROBABILIDAD · Exposición de la empresa (peso 50 %)", me.EXPOSICION),
        ("PROBABILIDAD · Frecuencia del delito (peso 30 %) — INE, Estadística "
         "de Condenados · CGPJ", me.FRECUENCIA),
        ("PROBABILIDAD · Historial de la empresa (peso 20 %)", me.HISTORIAL),
        ("IMPACTO · Severidad de la pena prevista (art. 33.7 CP)", me.SEVERIDAD),
    ]
    for titulo, tabla in bloques:
        ws.cell(fila, 1, titulo).font = Font(bold=True, size=11, color=ROJO)
        fila += 1
        _cabecera(ws, fila, ["Valor", "Criterio"], [10, 118])
        fila += 1
        for k in sorted(tabla):
            ws.cell(fila, 1, k).alignment = Alignment(horizontal="center")
            ws.cell(fila, 2, tabla[k]).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[fila].height = 26
            fila += 1
        fila += 1
    ws.cell(fila, 1, "IMPACTO · Ajuste según lo que expone ESTA empresa").font = \
        Font(bold=True, size=11, color=ROJO)
    fila += 1
    _cabecera(ws, fila, ["Valor", "Criterio"], [10, 118])
    fila += 1
    for k in (-1, 0, 1):
        ws.cell(fila, 1, k).alignment = Alignment(horizontal="center")
        ws.cell(fila, 2, me.EXPUESTO[k]).alignment = Alignment(wrap_text=True)
        fila += 1
    fila += 2
    ws.cell(fila, 1, "LAS DOS REGLAS QUE PESAN MÁS QUE LA MEDIA").font = \
        Font(bold=True, size=11, color=ROJO)
    for t in ("Si la empresa NO realiza la actividad (exposición 1), la "
              "probabilidad es 1 y el delito se marca como no aplicable con "
              "su motivo escrito.",
              "Si YA la han sancionado en ese ámbito (historial 4 o 5), la "
              "probabilidad no baja de 4. Nadie puede sostener que es "
              "improbable aquello por lo que ya le sancionaron."):
        fila += 1
        ws.cell(fila, 1, t).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila, end_column=2)
        ws.row_dimensions[fila].height = 30
    fila += 3
    ws.cell(fila, 1, "BANDAS").font = Font(bold=True, size=11, color=ROJO)
    fila += 1
    _cabecera(ws, fila, ["Desde", "Hasta", "Nivel", "Tratamiento"],
              [10, 10, 16, 92])
    for desde, hasta, nombre, trato in me.BANDAS:
        fila += 1
        ws.cell(fila, 1, desde)
        ws.cell(fila, 2, hasta)
        ws.cell(fila, 3, nombre).font = Font(bold=True)
        ws.cell(fila, 4, trato)
    return ws


def _hoja_listas(wb):
    """La hoja maestra: listas desplegables y factores de eficacia.

    Aquí está la diferencia con la matriz vieja: allí el estado del control y
    su reducción eran dos listas independientes que nada unía, así que un
    control «comunicado» descontaba lo mismo que uno «implantado». Aquí el
    factor SALE del estado. No se puede desacoplar a mano.
    """
    ws = wb.create_sheet("LISTAS")
    ws.sheet_state = "visible"
    _titulo(ws, "Listas y factores", "No se escribe aquí a mano.")
    ws["A4"] = "Escala"
    ws["A4"].font = Font(bold=True)
    for i, v in enumerate((1, 2, 3, 4, 5), start=5):
        ws.cell(i, 1, v)
    ws["B4"] = "Ajuste"
    ws["B4"].font = Font(bold=True)
    for i, v in enumerate((-1, 0, 1), start=5):
        ws.cell(i, 2, v)
    ws["C4"] = "Aplica"
    ws["C4"].font = Font(bold=True)
    ws["C5"], ws["C6"] = "SÍ", "NO"

    pares = [("E", "Naturaleza", me.NATURALEZA),
             ("G", "Estado", me.ESTADO),
             ("I", "Evidencia", me.EVIDENCIA),
             ("K", "Origen", me.ORIGEN)]
    for col, nombre, tabla in pares:
        ws[f"{col}4"] = nombre
        ws[f"{col}4"].font = Font(bold=True)
        sig = get_column_letter(ws[f"{col}4"].column + 1)
        ws[f"{sig}4"] = "Factor"
        ws[f"{sig}4"].font = Font(bold=True)
        for i, (k, v) in enumerate(tabla.items(), start=5):
            ws.cell(i, ws[f"{col}4"].column, k)
            ws.cell(i, ws[f"{col}4"].column + 1, v)
    ws["M4"] = "Eficacia base"
    ws["M4"].font = Font(bold=True)
    ws["M5"] = me.EFICACIA_BASE
    ws["N4"] = "Suelo residual"
    ws["N4"].font = Font(bold=True)
    ws["N5"] = me.SUELO_RESIDUAL
    for col, ancho in (("A", 9), ("B", 9), ("C", 9), ("E", 14), ("F", 9),
                       ("G", 14), ("H", 9), ("I", 14), ("J", 9), ("K", 14),
                       ("L", 9), ("M", 14), ("N", 14)):
        ws.column_dimensions[col].width = ancho
    return ws


def _hoja_catalogo(wb):
    ws = wb.create_sheet("CATÁLOGO")
    _titulo(ws, f"Catálogo de delitos · versión {cat.VERSION}",
            "Art. 31 bis CP. Considerar TODOS los delitos es un requisito del "
            "apartado 4.5.2 de la UNE 19601:2025: descartar es escribir por qué.")
    _cabecera(ws, 4, ["ID", "Delito o familia", "Familia", "Referencia",
                      "Severidad propuesta", "Nota"],
              [8, 52, 26, 30, 12, 46])
    for i, (idd, nombre, fam, ref, sev, nota) in enumerate(cat.DELITOS, start=5):
        ws.cell(i, 1, idd)
        ws.cell(i, 2, nombre)
        ws.cell(i, 3, fam)
        ws.cell(i, 4, ref)
        ws.cell(i, 5, sev).alignment = Alignment(horizontal="center")
        c = ws.cell(i, 6, nota)
        if "verificar" in (nota or "").lower():
            c.font = Font(color="B23B3B", bold=True)
        if "RECUPERADO" in (nota or ""):
            c.font = Font(color="0F5F3A", bold=True)
    fila = len(cat.DELITOS) + 7
    ws.cell(fila, 1, "FUERA DEL ART. 31 BIS · consecuencias accesorias del "
                     "art. 129: otro régimen, no se mezclan aquí").font = \
        Font(bold=True, color=ROJO)
    for j, (idd, nombre, ref) in enumerate(cat.ART_129, start=fila + 1):
        ws.cell(j, 1, idd)
        ws.cell(j, 2, nombre)
        ws.cell(j, 4, ref)
    return ws


def _hoja_controles(wb):
    """Los controles, con su eficacia calculada por fórmula."""
    ws = wb.create_sheet("CONTROLES")
    _titulo(ws, "Controles",
            "La eficacia SALE del estado, de la naturaleza, de la evidencia y "
            "del origen. No se pone a mano.")
    _cabecera(ws, 4, ["Nº", "Descripción del control", "Conducta que reduce",
                      "Naturaleza", "Estado", "Evidencia", "Origen",
                      "Documento que lo acredita", "Eficacia", "ln(1-ef)"],
              [6, 54, 18, 15, 15, 15, 15, 34, 11, 11])
    ini, fin = 5, 4 + FILAS_CONTROLES
    for r in range(ini, fin + 1):
        ws.cell(r, 1, r - 4)
        # La eficacia: base x los cuatro factores, cada uno buscado en LISTAS.
        # Un control IGNORADO da factor 0 y desaparece del cálculo: no es un
        # control, es una intención.
        ws.cell(r, 9).value = (
            f'=IF(OR($C{r}="",$D{r}="",$E{r}=""),"",'
            f'LISTAS!$M$5'
            f'*IFERROR(LOOKUP($D{r},LISTAS!$E$5:$E$7,LISTAS!$F$5:$F$7),0)'
            f'*IFERROR(LOOKUP($E{r},LISTAS!$G$5:$G$8,LISTAS!$H$5:$H$8),0)'
            f'*IFERROR(LOOKUP($F{r},LISTAS!$I$5:$I$6,LISTAS!$J$5:$J$6),0)'
            f'*IFERROR(LOOKUP($G{r},LISTAS!$K$5:$K$6,LISTAS!$L$5:$L$6),0))')
        ws.cell(r, 9).number_format = "0%"
        # EL TRUCO QUE HACE MULTIPLICATIVO UN SUMAR.SI: el producto de
        # (1-eficacia) es la exponencial de la suma de sus logaritmos. Así se
        # conserva la función que ya usaban y se acumula multiplicando.
        ws.cell(r, 10).value = f'=IF(N($I{r})=0,0,LN(1-$I{r}))'
        ws.cell(r, 10).number_format = "0.0000"
        ws.cell(r, 10).font = Font(color="B9B3A8", size=9)
    for col, origen, ancho in (("D", "$E$5:$E$7", 3), ("E", "$G$5:$G$8", 4),
                               ("F", "$I$5:$I$6", 2), ("G", "$K$5:$K$6", 2)):
        dv = DataValidation(type="list",
                            formula1=f"=LISTAS!{origen}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}{ini}:{col}{fin}")
    return ws


def _hoja_matriz(wb):
    """La matriz: una fila por conducta. Es donde vive el trabajo."""
    ws = wb.create_sheet("MATRIZ", 0)
    _titulo(ws, "Matriz de riesgos penales",
            "Una fila por CONDUCTA, no por delito: nadie controla «el "
            "blanqueo», se controla una conducta concreta.")
    _cabecera(ws, 4, [
        "Código", "Delito", "Conducta", "Área", "Dueño del riesgo",
        "Aplica", "Motivo si no aplica",
        "Exposición", "Frecuencia", "Historial", "PROBABILIDAD",
        "Severidad", "Ajuste empresa", "IMPACTO",
        "INHERENTE", "Nivel inh.", "Factor controles", "RESIDUAL", "Nivel res.",
        "Justificación de la puntuación", "Evidencia / fuente"],
        [10, 30, 46, 16, 18, 8, 30, 11, 11, 11, 13, 11, 13, 11, 11, 12, 13,
         11, 12, 46, 34])
    ini, fin = 5, 4 + FILAS_MATRIZ
    for r in range(ini, fin + 1):
        # PROBABILIDAD, con las dos reglas dentro de la propia fórmula para
        # que no dependan de que alguien se acuerde.
        ws.cell(r, 11).value = (
            f'=IF($F{r}<>"SÍ","",'
            f'IF($H{r}=1,1,'
            f'IF($J{r}>=4,MAX(4,ROUND(0.5*$H{r}+0.3*$I{r}+0.2*$J{r},0)),'
            f'ROUND(0.5*$H{r}+0.3*$I{r}+0.2*$J{r},0))))')
        ws.cell(r, 14).value = (
            f'=IF($F{r}<>"SÍ","",MEDIAN(1,$L{r}+N($M{r}),5))')
        ws.cell(r, 15).value = f'=IF($K{r}="","",$K{r}*$N{r})'
        ws.cell(r, 16).value = (
            f'=IF($O{r}="","",LOOKUP($O{r},{{1;5;10;16}},'
            f'{{"Bajo";"Medio";"Alto";"Crítico"}}))')
        # EL RANGO ES DE COLUMNA ENTERA A PROPÓSITO. En la matriz anterior el
        # SUMIF tenía el rango escrito a mano y nunca se amplió: 93 controles
        # de 603 no descontaban nada. Con la columna entera, añadir filas de
        # control no rompe jamás el cálculo.
        ws.cell(r, 17).value = (
            f'=IF($A{r}="","",EXP(SUMIF(CONTROLES!$C:$C,$A{r},CONTROLES!$J:$J)))')
        ws.cell(r, 17).number_format = "0%"
        ws.cell(r, 18).value = (
            f'=IF($O{r}="","",MAX($O{r}*$Q{r},$O{r}*LISTAS!$N$5))')
        ws.cell(r, 18).number_format = "0.0"
        ws.cell(r, 19).value = (
            f'=IF($R{r}="","",LOOKUP($R{r},{{1;5;10;16}},'
            f'{{"Bajo";"Medio";"Alto";"Crítico"}}))')
        for col in (11, 14, 15, 18):
            ws.cell(r, col).font = Font(bold=True)
    dv = DataValidation(type="list", formula1="=LISTAS!$C$5:$C$6",
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F{ini}:F{fin}")
    for col in ("H", "I", "J", "L"):
        d = DataValidation(type="list", formula1="=LISTAS!$A$5:$A$9",
                           allow_blank=True,
                           error="Valora entre 1 y 5 usando la hoja CRITERIOS")
        ws.add_data_validation(d)
        d.add(f"{col}{ini}:{col}{fin}")
    d = DataValidation(type="list", formula1="=LISTAS!$B$5:$B$7",
                       allow_blank=True)
    ws.add_data_validation(d)
    d.add(f"M{ini}:M{fin}")
    # Los colores de la banda, en el inherente y en el residual.
    for col in ("O", "R"):
        rango = f"{col}{ini}:{col}{fin}"
        ws.conditional_formatting.add(rango, CellIsRule(
            operator="between", formula=["16", "25"],
            fill=PatternFill("solid", fgColor="F5B7B1")))
        ws.conditional_formatting.add(rango, CellIsRule(
            operator="between", formula=["10", "15.999"],
            fill=PatternFill("solid", fgColor="FAD7A0")))
        ws.conditional_formatting.add(rango, CellIsRule(
            operator="between", formula=["5", "9.999"],
            fill=PatternFill("solid", fgColor="FCF3CF")))
        ws.conditional_formatting.add(rango, CellIsRule(
            operator="between", formula=["0.0001", "4.999"],
            fill=PatternFill("solid", fgColor="D5F5E3")))
    return ws


def _hoja_resumen(wb):
    """Distribución, no medias. Nadie gestiona un riesgo medio."""
    ws = wb.create_sheet("RESUMEN")
    _titulo(ws, "Resumen",
            "Cuántos riesgos hay en cada banda y cuántos bajan con los "
            "controles. La media de un riesgo no significa nada.")
    _cabecera(ws, 4, ["Banda", "Inherente", "Residual", "Bajan de banda"],
              [16, 14, 14, 16])
    for i, (_, _, nombre, _) in enumerate(me.BANDAS, start=5):
        ws.cell(i, 1, nombre).font = Font(bold=True)
        ws.cell(i, 2).value = f'=COUNTIF(MATRIZ!$P:$P,$A{i})'
        ws.cell(i, 3).value = f'=COUNTIF(MATRIZ!$S:$S,$A{i})'
        ws.cell(i, 4).value = f'=$B{i}-$C{i}'
    fila = 11
    ws.cell(fila, 1, "Conductas evaluadas").font = Font(bold=True)
    ws.cell(fila, 2).value = '=COUNTIF(MATRIZ!$F:$F,"SÍ")'
    ws.cell(fila + 1, 1, "Descartadas con motivo").font = Font(bold=True)
    ws.cell(fila + 1, 2).value = '=COUNTIF(MATRIZ!$F:$F,"NO")'
    ws.cell(fila + 2, 1, "Peor riesgo residual").font = Font(bold=True)
    ws.cell(fila + 2, 2).value = '=IFERROR(MAX(MATRIZ!$R:$R),"—")'
    ws.cell(fila + 3, 1, "Controles cargados").font = Font(bold=True)
    ws.cell(fila + 3, 2).value = '=COUNTA(CONTROLES!$B$5:$B$804)'
    ws.cell(fila + 4, 1, "Controles sin evidencia documental").font = \
        Font(bold=True, color="B23B3B")
    ws.cell(fila + 4, 2).value = \
        '=COUNTIFS(CONTROLES!$B$5:$B$804,"<>",CONTROLES!$F$5:$F$804,"<>documentada")'
    ws.cell(fila + 6, 1,
            "APETITO DE RIESGO APROBADO POR EL ÓRGANO DE GOBIERNO:").font = \
        Font(bold=True, color=ROJO)
    ws.cell(fila + 7, 1, "Umbral")
    ws.cell(fila + 8, 1, "Aprobado por")
    ws.cell(fila + 9, 1, "Fecha")
    ws.cell(fila + 10, 1,
            "Mientras esté vacío, el informe no puede decir qué supera el "
            "apetito ni qué plan de acción es obligatorio. Esa ausencia es, "
            "en sí misma, un hallazgo.").font = Font(italic=True, size=9,
                                                     color="B23B3B")
    ws.cell(fila + 12, 1, "Por encima del umbral").font = Font(bold=True)
    ws.cell(fila + 12, 2).value = (
        f'=IF($B${fila+7}="","sin umbral aprobado",'
        f'COUNTIF(MATRIZ!$R:$R,">"&$B${fila+7}))')
    return ws


def _hoja_control_documental(wb, proyecto=None):
    """Lo que exige un sistema de calidad para que esto sea un documento."""
    ws = wb.create_sheet("CONTROL DOCUMENTAL")
    _titulo(ws, "Control documental",
            "ISO 9001 · información documentada. Un documento sin código, "
            "versión y firma no es un entregable.")
    p = proyecto or {}
    campos = [
        ("Código del documento", p.get("codigo", "")),
        ("Título", "Matriz de riesgos penales"),
        ("Empresa", p.get("empresa", "")),
        ("Versión", p.get("version", "1.0")),
        ("Estado", "borrador"),
        ("Elaborado por", p.get("elaborado", "")),
        ("Revisado por", p.get("revisado", "")),
        ("Aprobado por", p.get("aprobado", "")),
        ("Fecha de emisión", datetime.now().strftime("%d/%m/%Y")),
        ("Alcance", p.get("alcance", "")),
        ("Periodo", p.get("periodo", "")),
        ("Audiencia", p.get("audiencia", "")),
        ("Catálogo de delitos", f"versión {cat.VERSION}"),
        ("Cotejado con el CP el", p.get("catalogo_cotejado", "")),
        ("Próxima revisión", p.get("revision_proxima", "")),
    ]
    for i, (n, v) in enumerate(campos, start=4):
        ws.cell(i, 1, n).font = Font(bold=True)
        ws.cell(i, 2, v)
    fila = len(campos) + 6
    ws.cell(fila, 1, "DISPARADORES DE REVISIÓN · UNE 19601:2025, apartado 6.3 "
                     "planificación de los cambios").font = \
        Font(bold=True, color=ROJO)
    for j, t in enumerate((
            "Reforma del Código Penal que afecte al catálogo.",
            "Nueva actividad, nuevo producto o nuevo mercado.",
            "Nuevo centro de trabajo, filial o jurisdicción.",
            "Incidente, denuncia, inspección o expediente sancionador.",
            "Cambio del órgano de compliance penal o del órgano de gobierno.",
            "Resultado de auditoría interna o externa."), start=fila + 1):
        ws.cell(j, 1, "•")
        ws.cell(j, 2, t)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 82
    return ws


def _hoja_plan(wb):
    ws = wb.create_sheet("PLAN DE ACCIÓN")
    _titulo(ws, "Plan de acción",
            "Cada hallazgo con responsable, plazo y verificación de eficacia. "
            "Sin la verificación, una acción sólo demuestra que alguien hizo "
            "algo, no que sirviera.")
    _cabecera(ws, 4, ["Nº", "Conducta", "Clasificación", "Criterio",
                      "Evidencia", "Hecho", "Riesgo", "Acción",
                      "Responsable", "Plazo", "Verificación de eficacia",
                      "Verificada", "Cerrada"],
              [6, 30, 20, 24, 26, 34, 26, 40, 18, 12, 30, 11, 11])
    dv = DataValidation(
        type="list",
        formula1='"No conformidad mayor,No conformidad menor,Observación,'
                 'Oportunidad de mejora"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("C5:C300")
    for col in ("L", "M"):
        d = DataValidation(type="list", formula1='"SÍ,NO"', allow_blank=True)
        ws.add_data_validation(d)
        d.add(f"{col}5:{col}300")
    return ws


def _hoja_limitaciones(wb):
    ws = wb.create_sheet("LIMITACIONES")
    _titulo(ws, "Limitaciones y alcance",
            "Va siempre, aunque el resultado sea bueno.")
    textos = [
        "Este documento NO declara conformidad, certificabilidad ni "
        "adecuación plena a UNE 19601:2025 de ninguna organización.",
        "La probabilidad y el impacto se apoyan en la evidencia recibida y en "
        "el alcance revisado. La falta de evidencia no equivale a "
        "incumplimiento: se registra como evidencia pendiente.",
        "Los controles sin documento que los acredite figuran como "
        "«declarada» y su eficacia queda reducida. Eso no significa que no "
        "existan; significa que no se han podido verificar.",
        "El riesgo residual nunca es cero. El suelo del "
        f"{int(me.SUELO_RESIDUAL*100)} % del inherente es una decisión de "
        "método: ningún sistema de control elimina un riesgo.",
        "El catálogo de delitos debe cotejarse contra el Código Penal "
        "consolidado antes de emitir el informe. Ver la hoja CATÁLOGO: hay "
        "referencias marcadas para verificar.",
        "El apetito de riesgo lo aprueba el órgano de gobierno de la empresa. "
        "Elece no lo propone.",
    ]
    for i, t in enumerate(textos, start=4):
        c = ws.cell(i, 1, "• " + t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 42
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    ws.column_dimensions["A"].width = 24
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22
    return ws


# ---------------------------------------------------------------------------
def generar(destino, proyecto=None):
    """Escribe el .xlsx completo y devuelve la ruta."""
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_matriz(wb)
    _hoja_controles(wb)
    _hoja_resumen(wb)
    _hoja_criterios(wb)
    _hoja_catalogo(wb)
    _hoja_plan(wb)
    _hoja_control_documental(wb, proyecto)
    _hoja_limitaciones(wb)
    _hoja_listas(wb)
    wb.move_sheet("MATRIZ", offset=-9)
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino
